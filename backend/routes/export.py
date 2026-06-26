"""Case 7 — Excel export (by date range + filters) and PDF summary report."""
import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional

from core.db import db
from core.security import get_current_user
from core.utils import today_ist
from routes.leads import build_query, query_params_dep

router = APIRouter(prefix="/export", tags=["export"])


async def _label_maps():
    tags = {t["id"]: t["name"] for t in await db.catalogs.find({"type": "tag"}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)}
    users = {u["id"]: u["name"] for u in await db.users.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)}
    lost = {r["id"]: r["name"] for r in await db.catalogs.find({"type": "lost_reason"}, {"_id": 0, "id": 1, "name": 1}).to_list(200)}
    return tags, users, lost


COLUMNS = [
    ("id", "Lead ID"), ("contact_name", "Name"), ("phone", "Phone"), ("email_from", "Email"),
    ("city", "City"), ("state_name", "State"), ("lead_stage", "Lead Stage"), ("_tags", "Tags"),
    ("_agent", "Assigned Agent"), ("source_lead", "Source"), ("campaign_name", "Campaign"),
    ("ads_platform", "Ads Platform"), ("follow_up_tag", "Follow-up Tag"), ("follow_up_date", "Follow-up Date"),
    ("_status", "Status"), ("_lost", "Lost Reason"), ("create_date_ist", "Created (IST)"),
]


@router.get("/leads.xlsx")
async def export_leads_xlsx(params: dict = Depends(query_params_dep), user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    q = build_query(**params, current_user=user)
    tags, users, lost = await _label_maps()
    cursor = db.leads.find(q, {"_id": 0}).sort("create_date_ist", -1).limit(50000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    head_fill = PatternFill("solid", fgColor="4A90E2")
    head_font = Font(bold=True, color="FFFFFF")
    for ci, (_, header) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.fill = head_fill
        c.font = head_font
    r = 2
    async for lead in cursor:
        row = {
            **lead,
            "_tags": ", ".join(tags.get(t, str(t)) for t in (lead.get("tags") or [])),
            "_agent": users.get(lead.get("user_id"), ""),
            "_status": "Active" if lead.get("active", True) else "Lost/Archived",
            "_lost": lost.get(lead.get("lost_reason_id"), ""),
        }
        for ci, (key, _) in enumerate(COLUMNS, 1):
            val = row.get(key, "")
            ws.cell(row=r, column=ci, value=str(val) if val not in (None, False) else "")
        r += 1
    widths = [9, 22, 14, 26, 14, 14, 14, 24, 18, 16, 18, 13, 14, 16, 12, 18, 19]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"homeivf_leads_{today_ist()}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/report.pdf")
async def export_report_pdf(date_from: Optional[str] = None, date_to: Optional[str] = None,
                            user: dict = Depends(get_current_user)):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    flt = {"active": "all"}
    if date_from:
        flt["date_from"] = date_from
    if date_to:
        flt["date_to"] = date_to
    q = build_query(search=None, stage_id=None, lead_stage=None, tags=None, user_id=None,
                    source_lead=None, campaign_name=None, ads_platform=None, city=None, state_name=None,
                    active="all", date_from=date_from, date_to=date_to, follow_up=None, priority=None,
                    follow_up_tag=None, lost_reason_id=None, current_user=user)
    _, users, _ = await _label_maps()

    total = await db.leads.count_documents(q)
    converted = await db.leads.count_documents({**q, "lead_stage": "Converted"})

    async def group(field, label_map=None):
        rows = await db.leads.aggregate([
            {"$match": q}, {"$group": {"_id": f"${field}", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}, {"$limit": 25},
        ]).to_list(25)
        return [[label_map.get(r["_id"], r["_id"]) if label_map else (r["_id"] or "New / Unassigned"), r["count"]] for r in rows]

    by_stage = await group("lead_stage")
    by_source = await group("source_lead")
    by_agent = await group("user_id", users)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor("#357ABD"))
    story = [Paragraph("HomeIVF CRM — Lead Report", title)]
    rng = f"{date_from or 'beginning'} → {date_to or today_ist()}"
    story += [Paragraph(f"Date range: {rng}", styles["Normal"]),
              Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} · by {user['name']}", styles["Normal"]),
              Spacer(1, 8 * mm)]
    conv_rate = round(converted / total * 100, 1) if total else 0
    summ = Table([["Total Leads", "Converted", "Conversion Rate"], [str(total), str(converted), f"{conv_rate}%"]],
                 colWidths=[55 * mm, 55 * mm, 55 * mm])
    summ.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A90E2")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 1), (-1, 1), 16), ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))
    story += [summ, Spacer(1, 8 * mm)]

    def section(heading, rows):
        s = [Paragraph(f"<b>{heading}</b>", styles["Heading3"])]
        data = [["", "Count"]] + [[str(a), str(b)] for a, b in (rows or [["—", 0]])]
        t = Table(data, colWidths=[120 * mm, 45 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"), ("FONTSIZE", (0, 0), (-1, -1), 9)]))
        return s + [t, Spacer(1, 6 * mm)]

    story += section("Leads by Stage", by_stage)
    story += section("Leads by Source", by_source)
    story += section("Leads by Agent", by_agent)
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="homeivf_report_{today_ist()}.pdf"'})
